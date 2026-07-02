#!/usr/bin/env python3
"""
fill_tournees.py — injecte les soumissions journalières dans le fichier maître Excel.

Usage:
  python3 fill_tournees.py --master <path/to/master.xlsx> --submissions <fichier.json ou dossier/>

Le fichier --submissions peut être :
  - un fichier JSON contenant un tableau de soumissions  (format : data/tournees/submissions.json)
  - un dossier contenant des fichiers *.json individuels (un par soumission)

Pour chaque soumission, le script :
  1. Identifie la feuille correspondant à la date (nom de feuille contient la date en divers formats)
  2. Pour chaque ligne non vide (chauffeur ou véhicule renseigné) :
     - Trouve la ligne Excel dont une cellule correspond au N° de tournée ET une autre au client
     - Écrit CHAUFFEUR → colonne I (9), VÉHICULE → colonne J (10), OBSERVATIONS → colonne K (11)
     - Ne touche à aucune autre cellule (couleurs, formules, etc. préservés)
"""

import argparse
import json
import os
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print('[fill_tournees] ERREUR : openpyxl non installé. Lancez : pip3 install openpyxl', file=sys.stderr)
    sys.exit(1)

# Colonnes cibles (1-indexed, correspondant aux colonnes Excel I / J / K)
COL_CHAUFFEUR    = 9
COL_VEHICULE     = 10
COL_OBSERVATIONS = 11


def norm(s):
    """Normalisation pour la comparaison : strip, majuscules, espaces uniques."""
    return ' '.join(str(s or '').strip().upper().split())


def date_variants(date_iso):
    """
    Renvoie les variantes de représentation d'une date ISO pour matcher
    les différents formats possibles de noms de feuilles.
    """
    try:
        d = datetime.fromisoformat(date_iso)
    except Exception:
        return [date_iso]
    variants = [
        date_iso,               # 2026-07-03
        d.strftime('%d/%m/%Y'), # 03/07/2026
        d.strftime('%d/%m'),    # 03/07
        d.strftime('%d-%m-%Y'), # 03-07-2026
        d.strftime('%d'),       # 03
        str(d.day),             # 3
    ]
    # Variante sans zéro initial (Linux strftime %-d, compatible macOS aussi via lstrip)
    no_zero = f"{d.day}/{d.month}"  # 3/7
    if no_zero not in variants:
        variants.append(no_zero)
    return variants


def find_sheet(wb, date_iso):
    """
    Cherche la feuille Excel correspondant à la date donnée.
    Le nom de feuille peut contenir la date sous diverses formes (03/07, 2026-07-03, "Lundi 03/07"…).
    """
    variants = date_variants(date_iso)
    for sheet_name in wb.sheetnames:
        sn = sheet_name.strip()
        for v in variants:
            if v in sn:
                return wb[sheet_name]
    return None


def find_matching_row(ws, master_client, tournee):
    """
    Parcourt toutes les lignes de la feuille pour trouver celle où apparaissent
    à la fois le client ET le numéro de tournée (normalisés).
    Si aucune ligne ne contient les deux simultanément, retourne la première ligne
    contenant uniquement la tournée (repli).
    Retourne le numéro de ligne (1-indexed) ou None.
    """
    t_norm = norm(tournee)
    c_norm = norm(master_client)

    if not t_norm:
        return None

    fallback_row = None

    for row in ws.iter_rows():
        vals = set()
        for cell in row:
            if cell.value is not None:
                vals.add(norm(cell.value))
        if not vals:
            continue

        has_tournee = t_norm in vals
        has_client  = c_norm in vals

        if has_tournee and (has_client or not c_norm):
            return row[0].row
        if has_tournee and fallback_row is None:
            fallback_row = row[0].row

    return fallback_row


def load_submissions(path_str):
    """
    Charge les soumissions depuis un fichier JSON (tableau) ou un dossier (*.json).
    Trie par date + submitted_at pour un traitement chronologique.
    """
    p = Path(path_str)

    if p.is_file():
        try:
            data = json.loads(p.read_text('utf-8'))
            subs = data if isinstance(data, list) else [data]
        except Exception as e:
            print(f'[fill_tournees] ERREUR lecture {path_str} : {e}', file=sys.stderr)
            return []
    elif p.is_dir():
        subs = []
        for f in sorted(p.glob('*.json')):
            try:
                data = json.loads(f.read_text('utf-8'))
                if isinstance(data, list):
                    subs.extend(data)
                else:
                    subs.append(data)
            except Exception as e:
                print(f'[fill_tournees] ⚠ {f.name} ignoré : {e}', file=sys.stderr)
    else:
        print(f'[fill_tournees] ⚠ Chemin introuvable : {path_str}', file=sys.stderr)
        return []

    # Tri chronologique (les soumissions récentes écrasent les anciennes pour une même cellule)
    subs.sort(key=lambda s: (s.get('date', ''), s.get('submitted_at', '')))
    return subs


def main():
    ap = argparse.ArgumentParser(description='Injecte les soumissions tournées dans le fichier maître xlsx.')
    ap.add_argument('--master',      required=True, help='Chemin vers le fichier maître .xlsx')
    ap.add_argument('--submissions', required=True, help='Fichier JSON ou dossier de soumissions')
    args = ap.parse_args()

    if not os.path.exists(args.master):
        print(f'[fill_tournees] ERREUR : fichier maître introuvable : {args.master}', file=sys.stderr)
        sys.exit(1)

    subs = load_submissions(args.submissions)
    if not subs:
        print('[fill_tournees] Aucune soumission à traiter.')
        return

    print(f'[fill_tournees] Ouverture de {args.master} — {len(subs)} soumission(s)…')
    wb = openpyxl.load_workbook(args.master)

    written = 0
    skipped = 0

    for sub in subs:
        date = sub.get('date', '')
        ws   = find_sheet(wb, date)
        if ws is None:
            print(f'[fill_tournees] ⚠ Aucune feuille pour la date {date!r} '
                  f'(feuilles disponibles : {wb.sheetnames[:5]}…)', file=sys.stderr)
            skipped += 1
            continue

        for row_data in sub.get('rows', []):
            chauffeur    = (row_data.get('chauffeur',    '') or '').strip()
            vehicule     = (row_data.get('vehicule',     '') or '').strip()
            observations = (row_data.get('observations', '') or '').strip()

            # Ne pas écraser si la ligne est complètement vide
            if not chauffeur and not vehicule:
                continue

            client  = row_data.get('master_client', '') or row_data.get('site', '')
            tournee = row_data.get('tournee', '')

            row_idx = find_matching_row(ws, client, tournee)
            if row_idx is None:
                print(f'[fill_tournees] ⚠ Ligne non trouvée : '
                      f'date={date} client={client!r} tournée={tournee!r}', file=sys.stderr)
                skipped += 1
                continue

            ws.cell(row=row_idx, column=COL_CHAUFFEUR).value    = chauffeur    or None
            ws.cell(row=row_idx, column=COL_VEHICULE).value     = vehicule     or None
            ws.cell(row=row_idx, column=COL_OBSERVATIONS).value = observations or None

            print(f'[fill_tournees] ✓ {ws.title!r} L{row_idx} | {client} | {tournee} '
                  f'→ {chauffeur} / {vehicule}')
            written += 1

    if written > 0:
        wb.save(args.master)
        print(f'[fill_tournees] ✓ Sauvegardé — {written} ligne(s) écrite(s), {skipped} ignorée(s).')
    else:
        print(f'[fill_tournees] Aucune ligne écrite ({skipped} ignorée(s)).')


if __name__ == '__main__':
    main()
